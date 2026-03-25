from datetime import datetime, timedelta
from io import BytesIO
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.urls import reverse
from django.views.generic import TemplateView, View

from apps.core.decorators import company_required
from apps.core.permissions import can_edit_alert_settings, can_export_reporting_global_excel


from django.db.models import Sum
from django.utils import timezone

from apps.core.models import Company
from apps.accounting.models import EntryLine
from apps.invoicing.models import Invoice
from apps.treasury.models import BankAccount, CashForecastItem
from openpyxl import Workbook

from .forms import AlertConfigForm
from .models import AlertConfig


def _resolve_company_for_reporting(request):
    cabinet = getattr(request.user, "role", None) == "CABINET_ADMIN"
    company = getattr(request, "company", None)
    company_name = request.GET.get("company_name") or request.POST.get("company_name")
    company_id = request.GET.get("company_id") or request.POST.get("company_id")
    if cabinet and company_name:
        company = Company.objects.filter(is_active=True, nom__iexact=company_name.strip()).first()
    elif cabinet and company_id:
        company = Company.objects.filter(pk=company_id).first()
    return company, cabinet


@method_decorator(company_required, name="dispatch")
class AnalyticsView(LoginRequiredMixin, TemplateView):
    template_name = "reporting/analytics.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        company, cabinet = _resolve_company_for_reporting(self.request)

        now = timezone.now().date()
        months = []
        # 6 derniers mois incluant le mois courant
        y = now.year
        m = now.month
        for i in range(5, -1, -1):
            yy = y
            mm = m - i
            while mm <= 0:
                mm += 12
                yy -= 1
            months.append((yy, mm))

        ca_labels = [f"{yy}-{mm:02d}" for yy, mm in months]
        ca_data = []
        for yy, mm in months:
            start = datetime(yy, mm, 1).date()
            end = (start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            qs = Invoice.objects.all()
            if not cabinet:
                qs = qs.filter(company=company)
            elif company:
                qs = qs.filter(company=company)
            ca = qs.filter(date__gte=start, date__lte=end).aggregate(total=Sum("total_ht")).get("total") or 0
            ca_data.append(float(ca))

        # Solde trésorerie (actuel)
        bank_qs = BankAccount.objects.all()
        if not cabinet:
            bank_qs = bank_qs.filter(company=company)
        elif company:
            bank_qs = bank_qs.filter(company=company)
        solde = bank_qs.aggregate(total=Sum("current_balance")).get("total") or 0

        # Top clients (sur 30 jours)
        top_clients = []
        inv_scope = Invoice.objects.all()
        if not cabinet:
            inv_scope = inv_scope.filter(company=company)
        elif company:
            inv_scope = inv_scope.filter(company=company)
        recent_start = now - timedelta(days=30)
        # MVP : on n'utilise pas annotate complexe, on fait un calcul simple
        clients = list(inv_scope.filter(date__gte=recent_start).values_list("client_name", flat=True).distinct())
        for cname in clients[:5]:
            total = inv_scope.filter(client_name=cname).aggregate(total=Sum("total_ttc")).get("total") or 0
            top_clients.append((cname, float(total)))
        top_client_names = [c for c, _ in top_clients]
        top_client_values = [v for _, v in top_clients]

        # Charges/Produits (sur 30 jours)
        start_period = now - timedelta(days=30)
        lines_qs = EntryLine.objects.select_related("account", "entry")
        if not cabinet:
            lines_qs = lines_qs.filter(entry__company=company)
        elif company:
            lines_qs = lines_qs.filter(entry__company=company)
        lines_qs = lines_qs.filter(entry__date__gte=start_period, entry__date__lte=now, entry__validated=True)

        # Simplification : on recalc via net debit-credit selon type (MVP)
        charges_val = 0
        produits_val = 0
        for line in lines_qs:
            net = (line.debit or 0) - (line.credit or 0)
            if line.account.type == "CHARGE":
                charges_val += float(net)
            if line.account.type == "PRODUIT":
                produits_val += float(net)

        # Délais de paiement (MVP : moyenne de jours de retard sur impayées)
        overdue_qs = Invoice.objects.all()
        if not cabinet:
            overdue_qs = overdue_qs.filter(company=company)
        elif company:
            overdue_qs = overdue_qs.filter(company=company)
        overdue = overdue_qs.exclude(status="PAID").exclude(status="CANCELLED").filter(due_date__lt=now)
        delays = []
        for inv in overdue[:50]:
            delays.append((now - inv.due_date).days)
        avg_delay = sum(delays) / len(delays) if delays else 0

        ctx.update(
            {
                "ca_labels": ca_labels,
                "ca_data": ca_data,
                "solde": float(solde),
                "top_client_names": top_client_names,
                "top_client_values": top_client_values,
                "charges_val": charges_val,
                "produits_val": produits_val,
                "avg_delay": avg_delay,
                "company_name": self.request.GET.get("company_name"),
                "companies": Company.objects.filter(is_active=True).order_by("nom") if cabinet else None,
            }
        )
        return ctx


@method_decorator(company_required, name="dispatch")
class AlertsView(LoginRequiredMixin, TemplateView):
    template_name = "reporting/alerts.html"

    def get(self, request, *args, **kwargs):
        # Même logique via get_context_data pour simplifier
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if not can_edit_alert_settings(request.user):
            messages.error(request, "La configuration des alertes est réservée au gérant et au comptable.")
            return redirect("reporting:alerts")

        cabinet = getattr(request.user, "role", None) == "CABINET_ADMIN"
        company = getattr(request, "company", None)
        company_name = request.POST.get("company_name")
        company_id = request.POST.get("company_id")
        if cabinet and company_name:
            company = Company.objects.filter(is_active=True, nom__iexact=company_name.strip()).first()
        elif cabinet and company_id:
            company = Company.objects.filter(pk=company_id).first()

        if company is None:
            return redirect("reporting:alerts")

        form = AlertConfigForm(request.POST)
        if form.is_valid():
            cfg, _ = AlertConfig.objects.get_or_create(company=company, is_active=True)
            cfg.treasury_threshold = form.cleaned_data["treasury_threshold"]
            cfg.email_enabled = form.cleaned_data["email_enabled"]
            cfg.save()
            messages.success(request, "Alertes mises à jour.")
        if cabinet and company:
            return redirect(f"{reverse('reporting:alerts')}?{urlencode({'company_name': company.nom})}")
        return redirect("reporting:alerts")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        now = timezone.now().date()
        company, cabinet = _resolve_company_for_reporting(self.request)

        company_name = self.request.GET.get("company_name")

        # Si cabinet admin sans company sélectionnée, on fait une consolidation simple : threshold=0
        threshold = 0
        if company is not None:
            cfg = AlertConfig.objects.filter(company=company, is_active=True).first()
            if cfg:
                threshold = float(cfg.treasury_threshold)

        # Calcul trésorerie
        bank_qs = BankAccount.objects.all()
        if not cabinet:
            bank_qs = bank_qs.filter(company=company)
        elif company is not None:
            bank_qs = bank_qs.filter(company=company)

        solde = bank_qs.aggregate(total=Sum("current_balance")).get("total") or 0

        # Factures échues
        inv_qs = Invoice.objects.all()
        if not cabinet:
            inv_qs = inv_qs.filter(company=company)
        elif company is not None:
            inv_qs = inv_qs.filter(company=company)

        overdue_count = inv_qs.exclude(status="PAID").exclude(status="CANCELLED").filter(due_date__lt=now).count()

        # Écritures non validées depuis 7j
        from apps.accounting.models import AccountingEntry

        entry_qs = AccountingEntry.objects.all()
        if not cabinet:
            entry_qs = entry_qs.filter(company=company)
        elif company is not None:
            entry_qs = entry_qs.filter(company=company)

        unvalidated_count = entry_qs.filter(validated=False, created_at__lte=timezone.now() - timedelta(days=7)).count()

        alerts = []
        if float(solde) < float(threshold):
            alerts.append(f"Trésorerie sous le seuil ({solde:.2f} < {threshold:.2f}).")
        if overdue_count > 0:
            alerts.append(f"{overdue_count} facture(s) échue(s).")
        if unvalidated_count > 0:
            alerts.append(f"{unvalidated_count} écriture(s) non validée(s) depuis 7 jours.")

        ctx.update(
            {
                "alerts": alerts,
                "threshold": threshold,
                "company_name": company_name,
                "companies": Company.objects.filter(is_active=True).order_by("nom") if cabinet else None,
                "form": AlertConfigForm(initial={"treasury_threshold": threshold, "email_enabled": True}),
            }
        )
        return ctx


@method_decorator(company_required, name="dispatch")
class ReportingExcelExportView(LoginRequiredMixin, View):
    """
    Export Excel global reporting (MVP):
    - Ecritures (lignes)
    - Factures
    - Transactions bancaires
    - Previsions de tresorerie
    """

    def get(self, request):
        if not can_export_reporting_global_excel(request.user):
            messages.error(request, "Export global réservé au gérant et au comptable.")
            return redirect("reporting:analytics")

        from apps.accounting.models import AccountingEntry
        from apps.treasury.models import BankTransaction

        company, cabinet = _resolve_company_for_reporting(request)

        entries = AccountingEntry.objects.select_related("company", "created_by").order_by("-date", "-id")
        invoices = Invoice.objects.select_related("company").order_by("-date", "-id")
        transactions = BankTransaction.objects.select_related("bank_account", "bank_account__company").order_by("-date", "-id")
        forecasts = CashForecastItem.objects.select_related("company").order_by("-date", "-id")

        if not cabinet:
            entries = entries.filter(company=company)
            invoices = invoices.filter(company=company)
            transactions = transactions.filter(bank_account__company=company)
            forecasts = forecasts.filter(company=company)
        elif company is not None:
            entries = entries.filter(company=company)
            invoices = invoices.filter(company=company)
            transactions = transactions.filter(bank_account__company=company)
            forecasts = forecasts.filter(company=company)

        wb = Workbook()

        ws_entries = wb.active
        ws_entries.title = "Ecritures"
        ws_entries.append(["Company", "Date", "Reference", "Description", "Validee", "Creee par"])
        for e in entries[:5000]:
            ws_entries.append([
                getattr(e.company, "nom", ""),
                e.date.isoformat(),
                e.reference or "",
                e.description or "",
                "Oui" if e.validated else "Non",
                getattr(e.created_by, "email", ""),
            ])

        ws_invoices = wb.create_sheet("Factures")
        ws_invoices.append(["Company", "Numero", "Date", "Echeance", "Client", "Statut", "Total HT", "Total TTC"])
        for inv in invoices[:5000]:
            ws_invoices.append([
                getattr(inv.company, "nom", ""),
                inv.number,
                inv.date.isoformat(),
                inv.due_date.isoformat(),
                inv.client_name,
                inv.get_status_display(),
                float(inv.total_ht),
                float(inv.total_ttc),
            ])

        ws_transactions = wb.create_sheet("Transactions")
        ws_transactions.append(["Company", "Compte", "Date", "Description", "Type", "Montant", "Rapprochee"])
        for tx in transactions[:5000]:
            ws_transactions.append([
                getattr(tx.bank_account.company, "nom", ""),
                tx.bank_account.name,
                tx.date.isoformat(),
                tx.description or "",
                tx.transaction_type,
                float(tx.amount),
                "Oui" if tx.reconciled else "Non",
            ])

        ws_forecasts = wb.create_sheet("Previsions")
        ws_forecasts.append(["Company", "Date", "Description", "Type", "Montant", "Categorie", "Reel"])
        for f in forecasts[:5000]:
            ws_forecasts.append([
                getattr(f.company, "nom", ""),
                f.date.isoformat(),
                f.description,
                f.type,
                float(f.amount),
                f.category or "",
                "Oui" if f.is_actual else "Non",
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "reporting_global.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

